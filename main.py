import os
import glob
from datetime import datetime
import pandas as pd

from scraper.underdog import get_pickem_slate, print_pickem_summary
from scraper.vlr import (
    find_player_url,
    scrape_current_team,
    scrape_agent_stats_by_timespan,
    scrape_match_links,
    parse_match_page,
    group_kills_by_match,
    fetch_soup,
)

import openpyxl
from openpyxl.styles import PatternFill

MAX_MATCHES = 40
BASE_DATA_DIR = "data"
AGENT_DIR = os.path.join(BASE_DATA_DIR, "agent_stats")
KILLS_DIR = os.path.join(BASE_DATA_DIR, "kills_by_match")
os.makedirs(AGENT_DIR, exist_ok=True)
os.makedirs(KILLS_DIR, exist_ok=True)

def print_welcome():
    print("="*70)
    print("VALORANT Esports Pro Stats & Pick'em Line Analyzer")
    print("="*70)
    print("Choose an option:")
    print("  1. Search for specific player stats from vlr.gg")
    print("  2. See all Underdog Fantasy pick'em lines, with VLR stats")
    print()
    print("Data will be saved to the data/ folder.")
    print("="*70)
    print()

def pretty_player_info(player, team, url):
    print("-"*60)
    print(f"Player found: {player}")
    print(f"Current Team: {team if team else 'N/A'}")
    print(f"VLR Profile: {url}")
    print("-"*60)

def export_agent_stats(stats, player):
    filename = os.path.join(AGENT_DIR, f"{player}_agent_stats.xlsx")
    rows = []
    for span, agents in stats.items():
        for agent, vals in agents.items():
            row = {'Timespan': span, 'Agent': agent}
            row.update(vals)
            rows.append(row)
    pd.DataFrame(rows).to_excel(filename, index=False)
    print(f"[+] Agent stats exported for {player} → {filename}")

def export_kills_by_match(good_matches, player):
    filename = os.path.join(KILLS_DIR, f"{player}_kills_by_match.xlsx")
    summary = []
    detail = []
    for m in good_matches:
        summary.append({
            'Match': m['match'],
            'Date': m['date'],
            'Total Kills': m['total_kills'],
            'Maps Counted': len(m.get('map_kills', []))
        })
        for d in m.get('map_kills', []):
            detail.append({
                'Match': m['match'],
                'Map': d['map'],
                'Agent': d['agent'],
                'Kills': d['kills']
            })
    with pd.ExcelWriter(filename) as writer:
        pd.DataFrame(summary).to_excel(writer, sheet_name='Match Summary', index=False)
        pd.DataFrame(detail).to_excel(writer, sheet_name='Map Kills Detail', index=False)
    print(f"[+] Kills by match exported for {player} → {filename}")

def compute_averages(good_matches, windows=(5, 10, 25)):
    kills = [m['total_kills'] for m in good_matches]
    averages = {}
    for w in windows:
        averages[w] = round(sum(kills[:w]) / w, 2) if len(kills) >= w else None
    return averages

def color_code_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Find column indices for line and averages
    headers = {cell.value: i for i, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), 1)}
    col_line = headers.get("Line")
    avg_cols = [headers.get(f"AvgLast{n}") for n in [5, 10, 25] if headers.get(f"AvgLast{n}")]

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")     # Above
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")       # Below
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")    # Equal

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        line_cell = row[col_line - 1] if col_line else None
        try:
            line_val = float(line_cell.value) if line_cell and line_cell.value is not None else None
        except (TypeError, ValueError):
            line_val = None

        for col in avg_cols:
            avg_cell = row[col - 1] if col else None
            try:
                avg_val = float(avg_cell.value) if avg_cell and avg_cell.value is not None else None
            except (TypeError, ValueError):
                avg_val = None

            if avg_val is not None and line_val is not None:
                if avg_val > line_val:
                    avg_cell.fill = green
                elif avg_val < line_val:
                    avg_cell.fill = red
                else:
                    avg_cell.fill = yellow

    wb.save(file_path)
    print(f"[+] Excel color coding applied to: {file_path}")

def do_specific_player():
    player = input("Enter player name: ").strip()
    url = find_player_url(player)
    if not url:
        print(f"❌ Player '{player}' not found on VLR.gg")
        return

    soup = fetch_soup(url)
    team = scrape_current_team(soup) if soup else None
    pretty_player_info(player, team, url)

    print("[*] Scraping agent stats by timespan...")
    agent_stats = scrape_agent_stats_by_timespan(url)
    if agent_stats:
        print("[+] Agent stats scraped.")
        export_agent_stats(agent_stats, player)
    else:
        print("[!] No agent stats found.")

    print("[*] Scraping recent match links...")
    links = scrape_match_links(url)
    print(f"[*] Found {len(links)} matches. Scraping up to {MAX_MATCHES}...")

    all_maps = []
    for idx, link in enumerate(links[:MAX_MATCHES], 1):
        print(f"    [{idx}/{min(len(links), MAX_MATCHES)}] {link}")
        maps = parse_match_page(link, player)
        all_maps.extend(maps)

    print("[*] Filtering for matches with exactly 2 maps and >0 kills...")
    good_matches = group_kills_by_match(all_maps, player, max_maps=2)
    print(f"[+] {len(good_matches)} matches passed the filter.")

    export_kills_by_match(good_matches, player)
    avgs = compute_averages(good_matches)
    print("\nPlayer Summary:")
    print(f"  Current Team: {team if team else 'N/A'}")
    print(f"  Last 5 Match Kills (Map 1+2): {avgs[5]}")
    print(f"  Last 10 Match Kills (Map 1+2): {avgs[10]}")
    print(f"  Last 25 Match Kills (Map 1+2): {avgs[25]}")
    print("All stats exported to Excel (see data/ folder).")

def do_full_slate():
    slate = get_pickem_slate()
    if not slate:
        print("❌ No pick'em slate found (Underdog API may be down).")
        return

    print("\nUNDERDOG FANTASY PICK'EM LINES")
    print("-"*60)
    player_info = []
    for item in slate:
        if "Kills on Maps 1+2 O/U" in item["over_under"]["title"]:
            player = item["over_under"]["title"].replace(" Kills on Maps 1+2 O/U", "").strip()
            line = item.get('stat_value')
            player_info.append((player, line))
            print_pickem_summary(item)
    print("-"*60)
    print(f"[Info] Will fetch VLR stats for: {' | '.join([p for p, l in player_info])}")
    print("\n[*] Starting VLR stats scraping for each player...\n")

    summary_rows = []
    for idx, (player, line) in enumerate(player_info, 1):
        print(f"\n[{idx}/{len(player_info)}] ===== Processing player: {player} =====")
        url = find_player_url(player)
        if not url:
            print(f"    [Warn] No VLR.gg profile found for '{player}'. Skipping stats scraping.")
            summary_rows.append({'Player': player, 'Line': line, 'Team': None, 'AvgLast5': None, 'AvgLast10': None, 'AvgLast25': None})
            continue

        soup = fetch_soup(url)
        team = scrape_current_team(soup) if soup else None
        print(f"    [Team] {team if team else 'N/A'} | [VLR] {url}")

        agent_stats = scrape_agent_stats_by_timespan(url)
        if agent_stats:
            export_agent_stats(agent_stats, player)

        links = scrape_match_links(url)
        all_maps = []
        for i, link in enumerate(links[:MAX_MATCHES], 1):
            print(f"        [{i}/{min(len(links), MAX_MATCHES)}] {link}")
            maps = parse_match_page(link, player)
            all_maps.extend(maps)

        good_matches = group_kills_by_match(all_maps, player, max_maps=2)
        export_kills_by_match(good_matches, player)
        avgs = compute_averages(good_matches)
        summary_rows.append({
            'Player': player,
            'Line': line,
            'Team': team,
            'AvgLast5': avgs[5],
            'AvgLast10': avgs[10],
            'AvgLast25': avgs[25]
        })

        # Print colored summary for this player
        def color_val(val, line):
            if val is None or line is None:
                return f"\033[90mN/A\033[0m"
            try:
                val = float(val)
                line = float(line)
                if val > line:
                    return f"\033[92m{val}\033[0m"  # Green
                elif val < line:
                    return f"\033[91m{val}\033[0m"  # Red
                else:
                    return f"\033[93m{val}\033[0m"  # Yellow
            except Exception:
                return str(val)

        print(f"    {'Line':<10}: {line}")
        print(f"    {'Last 5':<10}: {color_val(avgs[5], line)}")
        print(f"    {'Last 10':<10}: {color_val(avgs[10], line)}")
        print(f"    {'Last 25':<10}: {color_val(avgs[25], line)}")

    print("\n[*] Exporting summary table for all players...")
    df_summary = pd.DataFrame(summary_rows, columns=['Player','Team','Line','AvgLast5','AvgLast10','AvgLast25'])
    file_base = f"underdog_slate_{datetime.now().strftime('%Y-%m-%d')}"
    pattern = os.path.join(BASE_DATA_DIR, f"{file_base}*.xlsx")
    existing = glob.glob(pattern)
    filename = f"{file_base}.xlsx" if not existing else f"{file_base}_{len(existing)+1}.xlsx"
    out_file = os.path.join(BASE_DATA_DIR, filename)
    df_summary.to_excel(out_file, index=False)
    color_code_excel(out_file)
    print(f"[Success] Pick'em summary table exported: {out_file}")

def main():
    print_welcome()
    print("1. Search for specific player stats from vlr.gg")
    print("2. See Underdog Fantasy pick'em lines (with VLR stats)")
    choice = input("Enter 1 or 2: ").strip()
    if choice == "1":
        do_specific_player()
    elif choice == "2":
        do_full_slate()
    else:
        print("Invalid input. Please enter '1' or '2'.")

if __name__ == '__main__':
    main()